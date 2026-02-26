"""
File service — handles parsing and storage of uploaded files.
Supported formats: PDF, CSV, XLSX, XLS, TXT
"""
import csv
import io
import json
import uuid
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.file import UploadedFile

MAX_FILE_SIZE = 10 * 1024 * 1024  # 10 MB
ALLOWED_EXTENSIONS = {"pdf", "csv", "xlsx", "xls", "txt"}
TABULAR_EXTENSIONS = {"csv", "xlsx", "xls"}


def _extension(filename: str) -> str:
    return filename.rsplit(".", 1)[-1].lower() if "." in filename else ""


# ── CSV separator detection ───────────────────────────────────────────────────

def detect_csv_separator(text: str) -> str:
    """Detect whether a CSV uses ',' or ';' as separator."""
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        return dialect.delimiter
    except csv.Error:
        # Fallback: count occurrences on the first line
        first_line = sample.split("\n")[0]
        return ";" if first_line.count(";") > first_line.count(",") else ","


def _decode_csv(content: bytes) -> str:
    """Decode CSV bytes, stripping UTF-8 BOM if present."""
    for encoding in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="replace")


def parse_csv_smart(text: str) -> tuple[str, list[list[str]]]:
    """
    Parse CSV text with auto-detected separator.
    Returns (separator, rows) where rows is a list of string lists.
    """
    sep = detect_csv_separator(text)
    reader = csv.reader(io.StringIO(text), delimiter=sep)
    rows = [row for row in reader if any(cell.strip() for cell in row)]
    return sep, rows


def parse_xlsx_to_grid(content: bytes) -> list[list[str]]:
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.worksheets[0]  # first sheet only for import
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(["" if c is None else str(c) for c in row])
    wb.close()
    # Remove trailing empty rows
    while rows and all(c == "" for c in rows[-1]):
        rows.pop()
    return rows


def parse_xls_to_grid(content: bytes) -> list[list[str]]:
    import xlrd

    wb = xlrd.open_workbook(file_contents=content)
    sheet = wb.sheets()[0]  # first sheet only for import
    rows = []
    for row_idx in range(sheet.nrows):
        rows.append([str(sheet.cell_value(row_idx, col)) for col in range(sheet.ncols)])
    return rows


# ── Text extraction (for AI context) ─────────────────────────────────────────

def _parse_txt(content: bytes) -> str:
    return content.decode("utf-8", errors="replace")


def _parse_pdf(content: bytes) -> str:
    import pypdf

    reader = pypdf.PdfReader(io.BytesIO(content))
    pages = []
    for page in reader.pages:
        text = page.extract_text()
        if text:
            pages.append(text.strip())
    return "\n\n".join(pages)


def _parse_xlsx_text(content: bytes) -> str:
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    lines = []
    for ws in wb.worksheets:
        lines.append(f"[Feuille: {ws.title}]")
        for row in ws.iter_rows(values_only=True):
            lines.append(",".join("" if c is None else str(c) for c in row))
    wb.close()
    return "\n".join(lines)


def _parse_xls_text(content: bytes) -> str:
    import xlrd

    wb = xlrd.open_workbook(file_contents=content)
    lines = []
    for sheet in wb.sheets():
        lines.append(f"[Feuille: {sheet.name}]")
        for row_idx in range(sheet.nrows):
            lines.append(
                ",".join(str(sheet.cell_value(row_idx, col)) for col in range(sheet.ncols))
            )
    return "\n".join(lines)


def extract_text(content: bytes, filename: str) -> str:
    """Parse a file and return its AI-readable text content."""
    ext = _extension(filename)
    if ext == "pdf":
        return _parse_pdf(content)
    if ext == "csv":
        # Store raw decoded text so separator detection works later
        return _decode_csv(content)
    if ext == "xlsx":
        return _parse_xlsx_text(content)
    if ext == "xls":
        return _parse_xls_text(content)
    if ext == "txt":
        return _parse_txt(content)
    return ""


def extract_structured(content: bytes, filename: str) -> str | None:
    """
    For tabular files: return a JSON-serialised list[list[str]] (2D grid).
    Returns None for non-tabular formats.
    """
    ext = _extension(filename)
    try:
        if ext == "csv":
            text = _decode_csv(content)
            _, rows = parse_csv_smart(text)
            return json.dumps(rows, ensure_ascii=False)
        if ext == "xlsx":
            rows = parse_xlsx_to_grid(content)
            return json.dumps(rows, ensure_ascii=False)
        if ext == "xls":
            rows = parse_xls_to_grid(content)
            return json.dumps(rows, ensure_ascii=False)
    except Exception:
        return None
    return None


def get_grid_from_file(f: UploadedFile) -> tuple[str | None, list[list[str]]]:
    """
    Return (separator, grid) from a stored UploadedFile.
    Separator is only meaningful for CSV files.
    """
    if f.structured_data:
        rows: list[list[str]] = json.loads(f.structured_data)
        sep = None
        if f.file_type == "csv" and f.extracted_text:
            sep = detect_csv_separator(f.extracted_text)
        return sep, rows

    # Fallback for files uploaded before structured_data was added (CSV only)
    if f.file_type == "csv" and f.extracted_text:
        sep, rows = parse_csv_smart(f.extracted_text)
        return sep, rows

    return None, []


# ── DB operations ─────────────────────────────────────────────────────────────

async def create_file(
    db: AsyncSession,
    user_id: uuid.UUID,
    original_name: str,
    file_type: str,
    file_size: int,
    extracted_text: str,
    structured_data: str | None = None,
) -> UploadedFile:
    f = UploadedFile(
        user_id=user_id,
        original_name=original_name,
        file_type=file_type,
        file_size=file_size,
        extracted_text=extracted_text,
        structured_data=structured_data,
    )
    db.add(f)
    await db.commit()
    await db.refresh(f)
    return f


async def list_files(db: AsyncSession, user_id: uuid.UUID) -> list[UploadedFile]:
    result = await db.execute(
        select(UploadedFile)
        .where(UploadedFile.user_id == user_id)
        .order_by(UploadedFile.created_at.desc())
    )
    return list(result.scalars().all())


async def get_file(
    db: AsyncSession, file_id: uuid.UUID, user_id: uuid.UUID
) -> UploadedFile | None:
    result = await db.execute(
        select(UploadedFile).where(
            UploadedFile.id == file_id,
            UploadedFile.user_id == user_id,
        )
    )
    return result.scalar_one_or_none()


async def delete_file(db: AsyncSession, file_id: uuid.UUID, user_id: uuid.UUID) -> bool:
    f = await get_file(db, file_id, user_id)
    if not f:
        return False
    await db.delete(f)
    await db.commit()
    return True
